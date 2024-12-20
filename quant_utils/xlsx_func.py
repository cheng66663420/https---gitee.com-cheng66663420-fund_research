import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class XlsxFunc:
    def __init__(self, file_path: str = None, sheet_name: str = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        # 如果不存在路径,则新建一个
        if self.file_path is not None and os.path.exists(self.file_path):
            self.wb = load_workbook(self.file_path)

        else:
            self.wb = Workbook()
            print(f"不存在{self.file_path}")
        # 如果sheet_name为空，则默认为第一个sheet
        if self.sheet_name is None:
            self.sheet_name = self.wb.sheetnames[0]
        else:
            self.sheet_name = sheet_name
            # 如果不在workbook中，则抛出异常
            if self.sheet_name not in self.wb.sheetnames:
                raise ValueError(f"Sheet name {self.sheet_name} not found in workbook")
        self.ws = self.wb[self.sheet_name]

    @property
    def max_row(self):
        return self.ws.max_row

    @property
    def max_column(self):
        return self.ws.max_column

    def get_cells(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
    ) -> tuple:
        """
        获取单元格

        Parameters
        ----------
        cell_string : str, optional
            单元格,例如"A1"或者"A1:C3",默认None
        row : int, optional
            行数,自1开始,默认None
        column : int, optional
            列数,自1开始,默认None
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        Returns
        -------
        tuple
            ((cell1, cell2, ...), (cell1, cell2, ...), ...)
        """
        if cell_string is not None:
            cell = self.ws[cell_string]
            if isinstance(cell, tuple):
                return cell
            return ((cell,),)

        if row is not None and column is not None:
            col = get_column_letter(column)
            if end_row is not None and end_column is not None:
                end_col = get_column_letter(end_column)
                cell_string = f"{col}{row}:{end_col}{end_row}"
            else:
                cell_string = f"{col}{row}"
            cell = self.ws[cell_string]
            if isinstance(cell, tuple):
                return cell
            return ((cell,),)

        raise ValueError(
            "At least one of cell_string or row and column must be specified"
        )

    def get_cells_value(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
    ) -> list:
        """
        获取单元格的值

        Parameters
        ----------
        cell_string : str, optional
            单元格,例如"A1"或者"A1:C3",默认None
        row : int, optional
            行数,自1开始,默认None
        column : int, optional
            列数,自1开始,默认None
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        Returns
        -------
            单元格的值,
            格式为[[cell1.value, cell2.value], [cell3.value, cell4.value],...]
        """
        cell_tuples = self.get_cells(cell_string, row, column, end_row, end_column)
        result_list = []
        for cell_tuple in cell_tuples:
            result_list.append([cell.value for cell in cell_tuple])
        return result_list

    def set_cell_attribute(
        self, cell_tuples: tuple, attribute_name: str, attribute_value
    ):
        """
        设置单元格的属性

        Parameters
        ----------
        cell_tuples : tuple
            ((cell1, cell2, ...), (cell1, cell2, ...), ...)
        attribute_name : str
            属性名
        attribute_value : str
            属性值
        """
        for cell_tuple in cell_tuples:
            for cell in cell_tuple:
                try:
                    setattr(cell, attribute_name, attribute_value)
                except AttributeError:
                    # 如果属性不存在，则添加属性
                    raise AttributeError(f"Attribute '{attribute_name}' does not exist")

    def set_alignment(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
        alignment: Alignment = Alignment(horizontal="center", vertical="center"),
    ):
        """
        设置单元格的对齐方式

        Parameters
        ----------
        cell_string : str, optional
            单元格,例如"A1",默认None
        row : int
            行数,自从1开始
        column : int
            列名,自1开始
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        alignment : Alignment, optional
            对齐方式,默认水平居中,垂直居中
            by default Alignment(horizontal="center", vertical="center")
        """
        cell_tuples = self.get_cells(cell_string, row, column, end_row, end_column)
        self.set_cell_attribute(cell_tuples, "alignment", alignment)

    def set_value(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        value: float | str = None,
    ):
        """
        设置单元格的值

        Parameters
        ----------
        cell_string : str, optional
            单元格,例如"A1",默认None
        row : int
            行,自1开始
        column : int
            列,自1开始
        value : float | str
            单元格的值,数字或者字符串
        """
        cell_tuples = self.get_cells(cell_string, row, column)
        self.set_cell_attribute(cell_tuples, "value", value)

    def set_font(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
        font: Font = Font(name="微软雅黑", bold=True, color="000000", size=12),
    ):
        """
        设置单元格的字体

        Parameters
        ----------
        cell_string : str, optional
            单元格,例如"A1",默认None
        row : int
            行数
        column : int
            列数
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        font : Font, optional
            字体的格式, 默认微软雅黑,加粗,黑色,字号12
            by default Font(
            name="微软雅黑", bold=True,
            color="000000", size=12
        )
        """
        cell_tuples = self.get_cells(cell_string, row, column, end_row, end_column)
        self.set_cell_attribute(cell_tuples, "font", font)

    def set_fill(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
        fill: PatternFill = PatternFill(fgColor="FFFFFF", fill_type="solid"),
    ):
        """
        设置背景填充

        Parameters
        ----------
        cell_string : str, optional
            单元格,例如"A1",默认None
        row : int
            行数
        column : int
            列数
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        fill : PatternFill, optional
            背景填充,,默认白色,
            by default PatternFill(fgColor="FFFFFF", fill_type="solid")
        """
        cell_tuples = self.get_cells(cell_string, row, column, end_row, end_column)
        self.set_cell_attribute(cell_tuples, "fill", fill)

    def merge_cells(
        self,
        range_string: str = None,
        start_row: int = None,
        start_column: int = None,
        end_row: int = None,
        end_column: int = None,
    ):
        """
        合并单元格

        Parameters
        ----------
        start_row : int
            开始的行数
        start_column : int
            开始的列
        end_row : int
            结束的行
        end_column : int
            结束的列
        """
        self.ws.merge_cells(
            range_string=range_string,
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )

    def set_border(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
        border: Border = Border(
            top=Side(border_style="thick", color="000000"),
            right=Side(border_style="thick", color="000000"),
            bottom=Side(border_style="thick", color="000000"),
            left=Side(border_style="thick", color="000000"),
        ),
    ):
        """
        设置 单元格的边框

        Parameters
        ----------
        cell_string : str, optional
            表格字符串例如"A1", by default None
        row : int, optional
            行数, by default None
        column : int, optional
            列数, by default None
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        border : Border, optional
            表格样式, by default Border( top=Side(border_style="thick", color="000000"), right=Side(border_style="thick", color="000000"), bottom=Side(border_style="thick", color="000000"), left=Side(border_style="thick", color="000000"), )
        """
        cell_tuples = self.get_cells(cell_string, row, column, end_row, end_column)
        self.set_cell_attribute(cell_tuples, "boder", border)

    def set_number_format(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
        number_format: str = "0.00",
    ):
        """
        设置数字格式

        Parameters
        ----------
        row : int
            行数
        column : int
            列数
        number_format : str, optional
            数字格式, 默认0.00,
            by default "0.00"
        """
        cell_tuples = self.get_cells(cell_string, row, column, end_row, end_column)
        self.set_cell_attribute(cell_tuples, "number_format", number_format)

    def clear_sheet(self):
        """
        清空表格
        """
        self.ws.delete_rows(1, self.ws.max_row)
        self.ws.delete_cols(1, self.ws.max_column)

    def delete_rows(self, row: int = None, end_row: int = None):
        self.ws.delete_rows(row, end_row)

    def delete_cols(self, column: int = None, end_column: int = None):
        self.ws.delete_cols(column, end_column)

    def unmerge_all_cells(self):
        """
        取消工作表中所有合并的单元格。

        :param worksheet: openpyxl工作表对象。
        """
        # 获取所有已合并的单元格范围
        merged_cells = list(self.ws.merged_cells.ranges)  # Create a list copy
        for merged_range in merged_cells:
            self.ws.unmerge_cells(str(merged_range))

    def format_painter(
        self,
        cell_string: str = None,
        row: int = None,
        column: int = None,
        end_row: int = None,
        end_column: int = None,
        source_cell_string: str = None,
        source_row: int = None,
        source_column: int = None,
        end_source_row: int = None,
        end_source_column: int = None,
    ):
        """
        格式复制,根据源目标单元格格式复制到目标单元格中
        注意
        Parameters
        ----------
        cell_string : str, optional
            表格字符串例如"A1", by default None
        row : int, optional
            行数, by default None
        column : int, optional
            列数, by default None
        end_row : int, optional
            结束行,自1开始,默认None
        end_column : int, optional
            结束列,自1开始,默认None
        soucer_cell_string : str, optional
            源表格字符串例如"A1", by default None
        soucer_row : int, optional
            源表格行数, by default None
        soucer_column : int, optional
            源表格列数, by default None
        end_soucer_row : int, optional
            源表格结束行,自1开始,默认None
        end_soucer_column : int, optional
            源表格结束列,自1开始,默认None
        """
        # 需要复制的单元格格式
        source_cell = self.get_cells(
            source_cell_string,
            source_row,
            source_column,
            end_source_row,
            end_source_column,
        )
        # 目标单元格格式设置
        target_cell = self.get_cells(cell_string, row, column, end_row, end_column)
        if len(source_cell) != len(target_cell):
            raise ValueError("源目标单元格与目标单元格区域大小不一致")
        if len(source_cell[0]) != len(target_cell[0]):
            raise ValueError("源目标单元格与目标单元格区域大小不一致")
        # 复制格式
        for idx1, val1 in enumerate(target_cell):
            for idx2, val2 in enumerate(val1):
                val2.font = Font(**source_cell[idx1][idx2].font.__dict__)
                val2.fill = PatternFill(**source_cell[idx1][idx2].fill.__dict__)
                val2.border = Border(**source_cell[idx1][idx2].border.__dict__)
                val2.number_format = source_cell[idx1][idx2].number_format
                val2.alignment = Alignment(**source_cell[idx1][idx2].alignment.__dict__)

    def save(self, new_path: str = None, new_sheet_name: str = None):
        """
        保存文件,如果new_path和new_sheet_name为None,则保存当前文件

        Parameters
        ----------
        new_path : str, optional
            保存路径,如果为空则在源文件上保存, by default None
        new_sheet_name : str, optional
            新的sheet名称, by default None
        """
        if new_path is None:
            new_path = self.file_path
        if new_sheet_name is None:
            new_sheet_name = self.sheet_name
        self.ws.title = new_sheet_name
        self.wb.save(new_path)
        self.close()

    def close(self):
        """
        关闭workbook
        """
        self.wb.close()

    def write_dataframe_into_xlsx(
        self,
        data_df: pd.DataFrame,
        is_write_header: bool = True,
        is_write_index: bool = True,
        start_row: int = 1,
        start_column: int = 1,
        if_clear_sheet: bool = True,
    ):
        """
        将pd.DataFrame写入xlsx文件中

        Parameters
        ----------
        data_df : pd.DataFrame
            需要写入的数据
        is_write_header : bool, optional
            是否写入表头, by default True
        is_write_index : bool, optional
            是否写入索引, by default True
        start_row : int, optional
            开始写入的行数, by default 1
        start_column : int, optional
            开始写入的列数, by default 1
        if_clear_sheet : bool, optional
            是否清空sheet, by default True
        """
        if if_clear_sheet:
            self.clear_sheet()

        if is_write_index:
            start_column += 1
            self.set_value(row=1, column=1, value=data_df.index.name)
            for i, idx in enumerate(data_df.index, start=2):
                self.set_value(row=i, column=1, value=idx)

        if is_write_header:
            for i, col in enumerate(data_df.columns, start=start_column):
                self.set_value(row=start_row, column=i, value=col)
            start_row += 1

        for i, row in enumerate(data_df.values, start=start_row):
            for j, col in enumerate(row, start=start_column):
                self.set_value(row=i, column=j, value=col)


if __name__ == "__main__":
    xlsx = XlsxFunc(file_path="asfasfa")
    print(xlsx.sheet_name)
